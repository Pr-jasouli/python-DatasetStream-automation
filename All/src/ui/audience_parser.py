import json
import logging
import os
import sys

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Protection, Alignment, Border
from openpyxl.utils.dataframe import dataframe_to_rows

from utilities.utils import show_message

logging.basicConfig(level=logging.DEBUG)


def load_excel_data(filepath):
    """Loads Excel data into a pandas DataFrame."""
    return pd.read_excel(filepath)


def get_column_names(df):
    """Dynamically fetches column names based on content criteria."""
    identifiers = ['DATA_TYPE', 'PROD_NUM', 'BUS_CHANL_NUM']
    viewing_columns = [col for col in df.columns if 'VIEWING' in col]
    return identifiers, viewing_columns


def duplicate_data(df, ref_year, ref_month, target_years, identifiers):
    """Duplicates data for specified target years resetting non-key columns."""
    months = range(1, int(ref_month) + 1)
    mask = (df['PERIOD_YEAR'] == int(ref_year)) & (df['PERIOD_MONTH'].isin(months))
    data_to_duplicate = df.loc[mask]

    zero_columns = [col for col in df.columns if col not in ['PERIOD_YEAR', 'PERIOD_MONTH'] + identifiers]
    new_rows = []
    for year in target_years:
        for _, row in data_to_duplicate.iterrows():
            new_row = row.copy()
            new_row['PERIOD_YEAR'] = year
            new_row.update({col: 0 for col in zero_columns})
            new_rows.append(new_row)

    if new_rows:
        new_rows_df = pd.DataFrame(new_rows)
        return pd.concat([df, new_rows_df], ignore_index=True)
    else:
        return df


def adjust_viewing_columns(df, ref_year, identifiers, viewing_columns):
    """
    Définition des colonnes somme : Les variables sum_eop_2024 et sum_eop_2025 représentent les noms des colonnes dans le DataFrame qui contiennent les valeurs à utiliser pour le calcul des ratios.
    """
    sum_eop_2024 = 'sum_eop_vol_2024'
    sum_eop_2025 = 'sum_eop_vol_2025'

    """
    Itération sur les lignes ciblées : La boucle itère sur toutes les lignes du DataFrame df où l'année (PERIOD_YEAR) est supérieure à l'année de référence (ref_year). Cela permet de cibler uniquement les lignes des années futures par rapport à l'année de référence.
    """
    for index, row in df[df['PERIOD_YEAR'] > int(ref_year)].iterrows():
        """
        Un masque est créé pour identifier la ligne de référence qui correspond au même mois (PERIOD_MONTH) et aux mêmes identifiants que la ligne actuelle traitée dans la boucle. Le masque est également conditionné à ce que l'année soit l'année de référence (ref_year).
        """
        ref_mask = (df['PERIOD_YEAR'] == int(ref_year)) & (df['PERIOD_MONTH'] == row['PERIOD_MONTH'])
        for id_col in identifiers:
            """
            Ce masque est affiné pour s'assurer qu'il correspond exactement aux identifiants de la ligne actuellement traitée.
            """
            ref_mask &= (df[id_col] == row[id_col])
        """
        Vérification de la présence de la ligne de référence : Avant de procéder à tout calcul, on vérifie si le DataFrame filtré par le masque n'est pas vide. Cela évite des erreurs lors de l'accès à un DataFrame vide.
        """
        if not df.loc[ref_mask].empty:
            ref_row = df.loc[ref_mask].iloc[0]
            """
            Si la ligne de référence existe et que la valeur dans sum_eop_2024 est supérieure à zéro, un ratio est calculé comme étant la valeur de sum_eop_2025 divisée par celle de sum_eop_2024.
            """
            if ref_row[sum_eop_2024] > 0:
                ratio = ref_row[sum_eop_2025] / ref_row[sum_eop_2024]
                """
                Ce ratio est ensuite utilisé pour ajuster les valeurs des colonnes spécifiées dans viewing_columns. Pour chaque colonne de viewing_columns, la nouvelle valeur est calculée en multipliant la valeur originale de la colonne par le ratio calculé.
                """
                for col in viewing_columns:
                    df.at[index, col] = ref_row[col] * ratio
                """
                Si la valeur dans sum_eop_2024 est zéro ou si la ligne de référence n'existe pas, les valeurs dans les colonnes de viewing_columns sont mises à zéro pour la ligne courante traitée.
                """
            else:
                df.loc[index, viewing_columns] = 0
        else:
            df.loc[index, viewing_columns] = 0


def copy_sheet(source_sheet, target_sheet):
    """Copy the content from the source sheet to the target sheet."""
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)

    for row_dim in source_sheet.row_dimensions.values():
        target_sheet.row_dimensions[row_dim.index].height = row_dim.height

    for col_dim in source_sheet.column_dimensions.values():
        target_sheet.column_dimensions[col_dim.index].width = col_dim.width

    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

def check_file_open(file_path):
    """Check if the file is open by trying to rename it."""
    if not os.path.isfile(file_path):
        return False
    try:
        os.rename(file_path, file_path)
    except OSError:
        return True
    return False


def save_dataframe_with_formatting(df, output_path, original_file, references_year):
    """Saves the modified DataFrame to a new sheet in the copied Excel file."""
    output_filepath = f"{output_path}\\forecast_audience.xlsx"

    if check_file_open(output_filepath):
        show_message("Error", f"The file {output_filepath} is open. Please close the file and try again.", type='error')
        return

    try:
        logging.info(f"Loading original workbook from {original_file}")
        workbook = load_workbook(original_file)
        reference_sheet = workbook.active

        new_reference_sheet = workbook.create_sheet(title="Reference (Copy)")
        copy_sheet(reference_sheet, new_reference_sheet)

        forecast_sheet = workbook.create_sheet(title="Forecast")

        future_df = df[df['PERIOD_YEAR'] > int(references_year)]

        logging.info("Writing data to the Forecast sheet")
        for r_idx, row in enumerate(dataframe_to_rows(future_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                forecast_sheet.cell(row=r_idx, column=c_idx, value=value)

        forecast_sheet.freeze_panes = 'A2'
        new_reference_sheet.freeze_panes = 'A2'

        forecast_sheet.auto_filter.ref = forecast_sheet.dimensions
        new_reference_sheet.auto_filter.ref = new_reference_sheet.dimensions

        logging.info("Adjusting column widths")
        for sheet in [forecast_sheet, new_reference_sheet]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        workbook.remove(reference_sheet)
        new_reference_sheet.title = "Reference"

        if 'Sheet1' in workbook.sheetnames:
            std = workbook['Sheet1']
            workbook.remove(std)

        logging.info(f"Saving workbook to {output_filepath}")
        workbook.save(output_filepath)
        logging.info(f"Data saved to {output_filepath}")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        show_message("Error", f"An error occurred: {e}", type='error')
        return


def main(args):
    params = json.loads(args)
    references_month = params['references_month']
    references_year = params['references_year']
    target_start_year = int(params['target_start_year'])
    target_end_year = int(params['target_end_year'])
    output_path = params['output_path']
    file_path = params['file_path']
    output_filepath = f"{output_path}\\forecast_audience.xlsx"

    if check_file_open(output_filepath):
        show_message("Error", f"The file {output_filepath} is open. Please close the file and try again.", type='error')
        return

    logging.info("Loading data from input file")
    df = load_excel_data(file_path)
    logging.info(f"DataFrame loaded with {len(df)} rows and {len(df.columns)} columns")
    identifiers, viewing_columns = get_column_names(df)
    target_years = range(target_start_year, target_end_year + 1)

    logging.info("Duplicating data for target years")
    df = duplicate_data(df, references_year, references_month, target_years, identifiers)
    logging.info("Adjusting viewing columns")
    adjust_viewing_columns(df, references_year, identifiers, viewing_columns)
    logging.info("Saving DataFrame with formatting")
    save_dataframe_with_formatting(df, output_path, file_path, references_year)


if __name__ == "__main__":
    main(sys.argv[1])
